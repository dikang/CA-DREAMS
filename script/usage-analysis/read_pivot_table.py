import openpyxl
import sys

vendors = {"Cadence": True, "Synopsys": True, "Ansys": True}
projects = {"CA DREAMS Hub Operations": True, "GaNAmP": True}
performers = {"MOSIS 2.0": True, "UCLA": True, "UCSD": True}
products = {"Virtuoso": True, "HSPICE": True, "HFSS": True}

if len(sys.argv) < 2:
    print("Usage: python match.py <pivot_table.xlsx> ")
    sys.exit(1)

file_a = sys.argv[1]
#file_b = sys.argv[2]
#file_c = sys.argv[3]

# Load workbook with data_only=True to get the displayed values (not formulas)
wb = openpyxl.load_workbook(file_a, data_only=True)
ws = wb["Sheet1"]

# Find the columns that correspond to "Row Labels" and "Sum of Total usage time (hours)"
header_row = None
headers = {}

for row in range(1, 11):  # scan rows 1–10
    cells = [c.value for c in ws[row] if c.value]
    if "Row Labels" in cells and "Sum of Total usage time (hours)" in cells:
        header_row = row
        # map header names → column indices
        for cell in ws[row]:
            if cell.value:
                headers[cell.value] = cell.column
        break

if not header_row:
    raise ValueError("Could not find a header row with expected pivot headers in rows 1–10.")

row_label_col = headers["Row Labels"]
usage_col = headers["Sum of Total usage time (hours)"]

if not (row_label_col and usage_col):
    raise ValueError("Couldn't find expected pivot headers in the first row.")

# Collect the pivot data (rows 2–10)
pivot_data = []
for row in range(header_row + 1, ws.max_row + 1):
    label = ws.cell(row=row, column=row_label_col).value
    usage = ws.cell(row=row, column=usage_col).value
    if label is None:
        continue  # skip empty rows
    pivot_data.append((label, usage))

# Print results
for label, usage in pivot_data:
    if label == None: 
        continue

    # --- Step 2.1: Determine category ---
    if label in projects:
        current_project = label
    elif label in performers:
        current_performer = label
    elif label in vendors:
        current_vendor = label
    elif label in products:
        current_product = label
    else:
        # Unknown label; skip or log
        continue

    # --- Step 2.2: If we have a valid time entry, record it ---
    if time_val is not None:
        key = {
            "project name": current_project,
            "performer": current_performer,
            "vendor name": current_vendor,
            "product name": current_product,
        }
        combined_data[f"{label}"] = {"info": key, "time": time_val}

# --- Step 3: Print results ---
for k, v in combined_data.items():
    print(f"{k} -> {v}")
